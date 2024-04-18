from PIL import Image
from PIL import ImageFont
from PIL import ImageDraw
import requests
from time import sleep
import json
import io
import os
import xlrd

import matplotlib.pyplot as plt
from matplotlib.patches import PathPatch
from matplotlib.offsetbox import OffsetImage, AnnotationBbox

import openpyxl

from pathlib import Path
savePath = str(Path().absolute()) + "\\archive\\"
storePath = str(Path().absolute()) + "\\pngs\\"

    
#########################################################################
def clear_pngs():
    files = os.listdir(storePath)
    for f in files:
        if f.endswith('.png'):
            os.remove(storePath + f)

def check_player(file_name, player_name, sheet_nm):
    wb = xlrd.open_workbook(file_name)
    sheet = wb.sheet_by_name(sheet_nm)
    for col in range(sheet.ncols):
        if player_name in sheet.cell_value(0, col):
            return True
    return False
    

def find_player_col(file_name, player_name, sheet_nm):
    wb = xlrd.open_workbook(file_name)
    sheet = wb.sheet_by_name(sheet_nm)
    for col in range(sheet.ncols):
        if sheet.cell_value(0,col) == '':
            break
        if player_name in sheet.cell_value(0,col):
            return col
    return -1

def make_pack(file_name, pack_no, col, output_name):
    wb = openpyxl.load_workbook(filename = file_name, data_only=True)
    sheet = wb['Draft']

    curr_row = 16 * (pack_no - 1) + 2
    max_row = 16 * pack_no + 1

    image_name = storePath +output_name +'.png'
    params = {'format' : 'json'}
    cards = []
    colors = []
    for row in range(curr_row, max_row):
        try:
            curr_card = sheet.cell(row, col).value
            color = sheet.cell(row,col).fill.start_color.rgb
            colors.append(color[2:])
            
        except Exception as e:
            print(e)
            break
        if curr_card == '':
            break
        response = requests.get('https://api.scryfall.com/cards/search?q=' + curr_card, params = params)
        card_dict = response.json()['data']
        for card_entry in card_dict:
            if card_entry['name'] == curr_card:
                print(curr_card)
                try:
                    temp_image_url = card_entry['image_uris']['normal']
                except:
                    temp_image_url = card_entry['card_faces'][0]['image_uris']['normal']
                temp_image = Image.open(io.BytesIO(requests.get(temp_image_url).content))
                cards.append(temp_image)
                break
        
    canvas_x = 2200
    canvas_y = 3080
    new_image = Image.new('RGB', (canvas_x,canvas_y))
    draw = ImageDraw.Draw(new_image)
    
    #draw BG
    curr_x = 0
    curr_y = 0
    card_count = 0
    for color in colors:
        draw.rectangle((curr_x,curr_y,curr_x+550,curr_y+770), fill='#'+color)
        curr_x += 550
        if curr_x >= canvas_x:
            curr_x = 0
            curr_y += 770
    #draw Cards
    curr_x = 31
    curr_y = 45
    for card in cards:
        new_image.paste(card, (curr_x,curr_y))
        curr_x += 550
        if curr_x >= canvas_x:
            curr_x = 31
            curr_y += 770
    #draw Legend
    myFont = ImageFont.truetype('888_MRG.ttf', 72)
    curr_x = 1680
    curr_y = 2355
    for row in sheet.iter_rows(min_row=1, max_row=1):
        for cell in row:
            if not cell or cell.value == '':
                break
            cell_color = cell.fill.start_color.rgb
            cell_color = cell_color[2:]
            draw.rectangle((curr_x, curr_y, curr_x +60, curr_y + 100), fill='#'+cell_color)
            try:
                draw.text((curr_x+75, curr_y), cell.value, (255,255,255), font=myFont)
                curr_y += 120
            except Exception as e:
                print(e)
    new_image = new_image.resize((round(new_image.size[0] *0.75), round(new_image.size[1]*0.75)))
    new_image.save(image_name)
